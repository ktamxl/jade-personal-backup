#!/usr/bin/env python3
import sys
sys.path.insert(0, '/workspace/pylibs')

import openpyxl
from datetime import datetime

INPUT_FILE = '/workspace/Excel_Fidelity_Holdings.xlsx'
OUTPUT_FILE = '/workspace/holdings_dump.txt'

wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)

lines = []
lines.append(f"=== Excel File: {INPUT_FILE} ===")
lines.append(f"Sheet count: {len(wb.sheetnames)}")
lines.append("")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"{'='*60}")
    lines.append(f"SHEET: {sheet_name}")
    lines.append(f"Dimensions: {ws.dimensions}  |  Max row: {ws.max_row}  |  Max col: {ws.max_column}")
    lines.append("")
    lines.append(f"{'ROW':<6} {'COL':<5} {'CELL REF':<10} {'VALUE'}")
    lines.append("-" * 80)

    row_count = 0
    for row in ws.iter_rows():
        row_vals = [cell.value for cell in row]
        # Only print non-empty rows
        if any(v is not None and str(v).strip() != '' for v in row_vals):
            row_count += 1
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != '':
                    val = cell.value
                    if isinstance(val, datetime):
                        val = val.strftime('%Y-%m-%d %H:%M:%S')
                    lines.append(f"  {cell.row:<4} {cell.column:<5} {str(cell.coordinate):<10} {repr(val)}")
    lines.append(f"\n  [Total non-empty rows in this sheet: {row_count}]")
    lines.append("")

lines.append(f"{'='*60}")
lines.append(f"DONE — Total sheets: {len(wb.sheetnames)}")

with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Output written to {OUTPUT_FILE}")
print(f"Total lines written: {len(lines)}")
